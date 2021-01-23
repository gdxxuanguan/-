from openpyxl import Workbook, load_workbook
def word_count_in_str(string, keyword):
    return len(string.split(keyword))-1
if __name__=='__main__':
    #初始化
    wb=load_workbook("筛选的微博分阶段评论.xlsx")
    wb2=load_workbook("积极消极词汇排名.xlsx")
    wb3=load_workbook("工作簿1.xlsx")

    sheet=wb.active
    sheetactive=wb2["积极词"]
    sheetnegative=wb2["消极词"]
    storeactive=wb3["Sheet1"]
    storenegative = wb3["Sheet2"]


    for hang in sheet[1]:
        i=65
        temple=chr(i)+"1"
        storeactive[temple]=hang.value
        storenegative[temple]=hang.value
        i=i+1
        if i==87:
            break

#积极词.
    count_set=[]

    for cell2 in sheetactive["A2:A100"]:
        count_set.append(cell2[0].value)

    for i in range(67,79):
        table=chr(i-2)
        for cell in sheet[table]:
            row = 2

            if cell.value is None:
                a=0
            else:
                value1=cell.value
                for item in count_set:
                    str1=chr(i)+str(row)
                    if storeactive[str1].value is None:
                        storeactive[str1] = word_count_in_str(value1, item)
                    else:
                        storeactive[str1]=storeactive[str1].value+word_count_in_str(value1,item)
                    row=row+1
                    print(row)
#消极词处理
    count_set = []

    for cell2 in sheetnegative["A2:A128"]:
        count_set.append(cell2[0].value)

    for i in range(67, 79):
        table = chr(i-2)
        for cell in sheet[table]:
            row = 2

            if cell.value is None:
                a = 0
            else:
                value1 = cell.value
                for item in count_set:
                    str1 = chr(i) + str(row)
                    if storenegative[str1].value is None:
                        storenegative[str1] = word_count_in_str(value1, item)
                    else:
                        storenegative[str1] = storenegative[str1].value + word_count_in_str(value1, item)
                    row = row + 1
                    print(row)

    wb3.save("工作簿1.xlsx")



