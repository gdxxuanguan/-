from openpyxl import Workbook, load_workbook
def word_count_in_str(string, keyword):
    return len(string.split(keyword))-1
i=1

wb3=load_workbook("工作簿1.xlsx")
for x in range(0,11):

    string1 = "NewsWeibo" + str(x) + ".xlsx"
    wb = load_workbook(string1)

    sheet=wb.active

    store=wb3["Sheet1"]

    count_set=["疫","医", "病毒", "新冠","诊","患","核酸","隔离","封城","疫苗","口罩","肺炎","传染","预防","居家","武汉","国家","英雄","物资","一线"]

    a=2
    for cell in sheet['D']:
        value1=cell.value
        fr1 = chr(65) + str(a)  # A
        fr2 = chr(66) + str(a)  # b
        fr3 = chr(67) + str(a) # c
        fr4 = chr(68) +str(a) # d
        fr5 = chr(69) + str(a)  # e
        fr6 = chr(70) + str(a) # f
        fr7 = chr(71) + str(a) # g
        fr8 = chr(72) + str(a) # h
        fr9 = chr(73) + str(a)  # i
        for item in count_set:
            if word_count_in_str(value1, item)>0:
                str1=chr(65)+str(i)#A
                str2 = chr(66) + str(i)#b
                str3 = chr(67) + str(i)#c
                str4 = chr(68) + str(i)#d
                str5 = chr(69) + str(i)#e
                str6 = chr(70) + str(i)#f
                str7 = chr(71) + str(i)#g
                str8 = chr(72) + str(i)#h
                str9 = chr(73) + str(i)#i
                k=chr(68)+str(i-1)
                if i==1:
                    store[str1] = sheet[fr1].value
                    store[str2] = sheet[fr2].value
                    store[str3] = sheet[fr3].value
                    store[str4] = sheet[fr4].value
                    store[str5] = sheet[fr5].value
                    store[str6] = sheet[fr6].value
                    store[str7] = sheet[fr7].value
                    store[str8] = sheet[fr8].value
                    store[str9] = sheet[fr9].value
                    i = i + 1
                else:
                    if not value1==store[k].value:

                        store[str1]=sheet[fr1].value
                        store[str2] = sheet[fr2].value
                        store[str3] = sheet[fr3].value
                        store[str4] = sheet[fr4].value
                        store[str5] = sheet[fr5].value
                        store[str6] = sheet[fr6].value
                        store[str7] = sheet[fr7].value
                        store[str8] = sheet[fr8].value
                        store[str9] = sheet[fr9].value
                        i=i+1
                break

        a=a+1
    print("已完成一个")
wb3.save("工作簿1.xlsx")
